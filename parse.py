import openpyxl, json, re, sys

SRC = '/mnt/user-data/uploads/SCHEDULE____202610.xlsx'
wb = openpyxl.load_workbook(SRC, data_only=True)

DAY_MAP = {'Lundi':'Mon','Mardi':'Tue','Mercredi':'Wed','Jeudi':'Thu','Vendredi':'Fri'}

def clean(v):
    if v is None: return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v

def parse_day(raw):
    if not raw: return None, None
    raw = str(raw).strip()
    # Reject descriptive/note cells (e.g. "1ère séance // Lundi 7 septembre...") -
    # a real day field starts with the day name, optionally followed only by a
    # "Semaines A/B" marker, with no other free text (no "//", no digits for dates).
    if '//' in raw or re.search(r'\d', raw):
        return None, None
    m = re.match(r'^(Lundi|Mardi|Mercredi|Jeudi|Vendredi)\b', raw)
    if not m: return None, None
    day = m.group(1)
    note = None
    if 'Semaines A' in raw: note = 'Semaine A'
    elif 'Semaines B' in raw: note = 'Semaine B'
    return DAY_MAP.get(day), note

def parse_times(raw):
    if not raw: return []
    raw = str(raw).strip()
    if raw in ('-', '', 'A préciser'): return []
    parts = [p.strip() for p in raw.split(chr(10)) if p.strip()]
    out = []
    for p in parts:
        m = re.match(r'(\d{3,4})\s*-\s*(\d{3,4})', p)
        if m:
            s, e = m.group(1), m.group(2)
            s = s.zfill(4); e = e.zfill(4)
            out.append((f'{s[:2]}:{s[2:]}', f'{e[:2]}:{e[2:]}'))
    return out

def build_schedule(day1_raw, time1_raw, day2_raw, time2_raw):
    schedule = []
    d1, note1 = parse_day(day1_raw)
    times1 = parse_times(time1_raw)
    if d1:
        for (s, e) in times1:
            schedule.append({'day': d1, 'start': s, 'end': e, 'note': note1})
    d2, note2 = parse_day(day2_raw)
    if d2:
        times2 = parse_times(time2_raw)
        for (s, e) in times2:
            schedule.append({'day': d2, 'start': s, 'end': e, 'note': note2})
    return schedule

def quad_group(type_str):
    if not type_str: return None
    m = re.search(r'Quadruplette\s+(\d+)', type_str)
    return int(m.group(1)) if m else None

courses = []
cid = 0

def next_id():
    global cid
    cid += 1
    return f'c{cid}'

REGULAR_SHEETS = {
    '1A - North America ': {'group': '1A_NA'},
    '1A -  Africa minor - English tr': {'group': '1A_AFRICA_EN'},
    '1A - Mineure Afrique & METIS ': {'group': '1A_AFRICA_FR'},
}

for sheet_name, meta in REGULAR_SHEETS.items():
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    has_day2 = len(header) > 10
    for row in rows[1:]:
        if not row[4]:
            continue
        programme, discipline, typ, up, title, code, ens1, ens2, jour1, horaire1 = row[:10]
        jour2 = row[10] if has_day2 and len(row) > 10 else None
        horaire2 = row[11] if has_day2 and len(row) > 11 else None
        schedule = build_schedule(jour1, horaire1, jour2, horaire2)
        courses.append({
            'id': next_id(),
            'group': meta['group'],
            'programme': clean(programme),
            'majeure': None,
            'discipline': clean(discipline),
            'type': clean(typ),
            'up': clean(up),
            'title': clean(title),
            'codeMatiere': clean(code),
            'teacher1': clean(ens1),
            'teacher2': clean(ens2),
            'schedule': schedule,
            'quadGroup': quad_group(typ),
            'langue': None,
            'niveau': None,
        })

SHEETS_2A = {
    '2A - North America & Africa min': {'group': '2A_NA'},
    '2A - Mineure Afrique & METIS': {'group': '2A_AFRICA_FR'},
}
for sheet_name, meta in SHEETS_2A.items():
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[1:]:
        if not row[5]:
            continue
        programme, majeure, discipline, typ, up, title, code, ens1, ens2, jour1, horaire1 = row[:11]
        schedule = build_schedule(jour1, horaire1, None, None)
        courses.append({
            'id': next_id(),
            'group': meta['group'],
            'programme': clean(programme),
            'majeure': clean(majeure),
            'discipline': clean(discipline),
            'type': clean(typ),
            'up': clean(up),
            'title': clean(title),
            'codeMatiere': clean(code),
            'teacher1': clean(ens1),
            'teacher2': clean(ens2),
            'schedule': schedule,
            'quadGroup': None,
            'langue': None,
            'niveau': None,
        })

ws = wb['2A - Séminaires ']
rows = list(ws.iter_rows(values_only=True))
for row in rows[1:]:
    if not row[2]:
        continue
    typ, up, title, code, ens1, ens2, jour1, horaire1 = row[:8]
    schedule = build_schedule(jour1, horaire1, None, None)
    courses.append({
        'id': next_id(),
        'group': '2A_SEMINAIRE',
        'programme': None,
        'majeure': 'Séminaire',
        'discipline': 'Séminaire',
        'type': clean(typ),
        'up': clean(up),
        'title': clean(title),
        'codeMatiere': clean(code),
        'teacher1': clean(ens1),
        'teacher2': clean(ens2),
        'schedule': schedule,
        'quadGroup': None,
        'langue': None,
        'niveau': None,
    })

ws = wb['Langues ']
rows = list(ws.iter_rows(values_only=True))
for row in rows[1:]:
    if not row[3]:
        continue
    langue, niveau, up, title, code, ens1, ens2, jour1, horaire1, jour2, horaire2 = row[:11]
    schedule = build_schedule(jour1, horaire1, jour2, horaire2)
    courses.append({
        'id': next_id(),
        'group': 'LANGUE',
        'programme': None,
        'majeure': None,
        'discipline': clean(langue),
        'type': 'Langue',
        'up': clean(up),
        'title': clean(title),
        'codeMatiere': clean(code),
        'teacher1': clean(ens1),
        'teacher2': clean(ens2),
        'schedule': schedule,
        'quadGroup': None,
        'langue': clean(langue),
        'niveau': clean(niveau),
    })

print('Total courses:', len(courses))
groups = {}
for c in courses:
    groups[c['group']] = groups.get(c['group'], 0) + 1
print(groups)

with open('/home/claude/build/courses.json', 'w', encoding='utf-8') as f:
    json.dump(courses, f, ensure_ascii=False, indent=1)
